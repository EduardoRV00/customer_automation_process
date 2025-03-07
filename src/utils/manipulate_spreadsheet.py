import pandas as pd
import logging
from config import *
from openpyxl import load_workbook
from src.tasks.processed_data import create_output_sheet


def access_spreadsheet_input(logger_client, logger_dev):
  """
  Loads the input spreadsheet and returns a dataFrame
  """
  try:
    df = pd.read_excel(EXCEL_RAW_PATH, dtype=str)
    print(f"Dados carregados da planilha:\n{df.head()}")
    logger_client.info("Planilha de entrada carregada com sucesso.")

    return df
  
  except Exception as e:
    logger_dev.error(f"Erro ao carregar a planilha de entrada: {e}")
    return None


def save_status_to_output(output_sheet, row_index, message, logger_client, logger_dev):
  """
  Saves a status message to a specific cell in the output sheet
  """
  try:
    logger_client.info(f"Acessando planilha de saída para salvar status.")
    wb = load_workbook(output_sheet)
    ws = wb.active

    status_column = 17
    status_cell = ws.cell(row=row_index + 2, column=status_column)

    #checks if there is already a value in the status field (new message added at the end)
    current_status = status_cell.value if status_cell.value else ""
    new_status = f"{current_status}, {message}".strip()

    status_cell.value = message
    wb.save(output_sheet)
    logger_client.info(f"Status salvo com sucesso na célula {status_cell.coordinate}")

  except Exception as e:
    logger_dev.error(f"Erro ao salvar status na linha {row_index + 2}: {e}")


def check_empty_fields(df, output_sheet, logger_client, logger_dev):
  """
  Check empty fields in the spreadsheet
  """
  fields = ["CNPJ", "VALOR DO PEDIDO", "DIMENSÕES CAIXA (altura x largura x comprimento cm)", "PESO DO PRODUTO", "TIPO DE SERVIÇO JADLOG", "TIPO DE SERVIÇO CORREIOS"]
  status_messages = []

  try:
    logger_client.info("Verifica campos vazios na planilha de entrada.")
    for index, row in df.iterrows():
      empty_fields = [field for field in fields if pd.isna(row[field]) or row[field] == ""]
      if empty_fields:
        message = f"Os campos {', '.join(empty_fields)} estão vazios"

        save_status_to_output(output_sheet, index, message, logger_client, logger_dev)
        status_messages.append((index, message))

  except Exception as e:
    logger_dev.error(f"Erro ao verificar campos vazios: {e}")

  return status_messages


def fill_output_sheet_with_input_data(input_df, output_sheet, logger_client, logger_dev):
  """
  Populates the output sheet with data from the input sheet, mapping specific columns.
  """
  try:
    logger_client.info("Preenchendo a planilha de saída com os dados da planilha de entrada...")
    wb = load_workbook(output_sheet)
    ws = wb.active

    for index, row in input_df.iterrows():
      cnpj = row.get('CNPJ', 'N/A')
      order_value = row.get('VALOR DO PEDIDO', 'N/A')
      box_dimension = row.get('DIMENSÕES CAIXA (altura x largura x comprimento cm)', 'N/A')
      product_weight = row.get('PESO DO PRODUTO', 'N/A')
      service_type_jadlog = row.get('TIPO DE SERVIÇO JADLOG', 'N/A')
      service_type_correios = row.get('TIPO DE SERVIÇO CORREIOS', 'N/A')

      ws.cell(row=index + 2, column=1).value = cnpj
      ws.cell(row=index + 2, column=9).value = order_value
      ws.cell(row=index + 2, column=10).value = box_dimension
      ws.cell(row=index + 2, column=11).value = product_weight
      ws.cell(row=index + 2, column=12).value = service_type_jadlog
      ws.cell(row=index + 2, column=13).value = service_type_correios

    wb.save(output_sheet)
    logger_client.info("Planilha de saída preenchida com sucesso!")
  except Exception as e:
    logger_dev.error(f"Erro ao preencher a planilha de saída: {e}") 


def process_spreadsheet(output_sheet, logger_client, logger_dev):
  """
  Access and check the spreadsheet
  """
  input_df = access_spreadsheet_input(logger_client, logger_dev)
  if input_df is not None:
    fill_output_sheet_with_input_data(input_df, output_sheet, logger_client, logger_dev)
    status_messages = check_empty_fields(input_df, output_sheet, logger_client, logger_dev)   
  return []
