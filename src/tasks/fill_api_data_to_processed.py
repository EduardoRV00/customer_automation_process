import pandas as pd
from openpyxl import load_workbook
from src.services.api_client import *
from src.utils.setup_logs import *
from src.utils.error_handling import *

def data_fill_processed(output_sheet, logger_client, logger_dev):
    """
    Fills the processed spreadsheet with data returned by the API and concatenates
    the API status message with any message already present in the cell.
    """
    try:
        msg_entry_sheet = "Lendo arquivo Excel de entrada"
        logger_client.info(msg_entry_sheet)
        logger_dev.info(msg_entry_sheet)
        entry_df = pd.read_excel(output_sheet)
        
        msg_processing_data = "Processando CNPJs e consultando API"
        logger_client.info(msg_processing_data)
        logger_dev.info(msg_processing_data)
        wb = load_workbook(output_sheet)
        ws = wb.active
        
        # Mapping columns where data will be inserted
        column_mapping = {
            "RAZÃO SOCIAL": 2,
            "NOME FANTASIA": 3,
            "ENDEREÇO": 4,
            "CEP": 5,
            "DESCRIÇÃO MATRIZ FILIAL": 6,
            "TELEFONE + DDD": 7,
            "E-MAIL": 8
        }
        
        msg_dict = "Armazenando informações em formato de Dict chave : valor"
        logger_client.info(msg_dict)
        logger_dev.info(msg_dict)
        for index, row in entry_df.iterrows():
            cnpj = str(row["CNPJ"])
            info = api_get(cnpj)
            
            if info is None:
                info = {
                    "razao_social": "Não disponível",
                    "nome_fantasia": "Não disponível",
                    "descricao_situacao_cadastral": "Não disponível",
                    "Endereco": "Não informado, S/N - Não informado",
                    "cep": "Não informado",
                    "identificador_matriz_filial": "Não informado",
                    "ddd_telefone_1": "Não informado",
                    "email": "N/A"
                }
            
            # Check if the company is active; if not, prepare the status message
            status_api = "" if info["descricao_situacao_cadastral"] == "ATIVA" else "Empresa inativa"
            
            # Fill data with correct cell info. 
            ws.cell(row=index + 2, column=column_mapping["RAZÃO SOCIAL"]).value = info["razao_social"]
            ws.cell(row=index + 2, column=column_mapping["NOME FANTASIA"]).value = info["nome_fantasia"]
            ws.cell(row=index + 2, column=column_mapping["ENDEREÇO"]).value = info["Endereco"]
            ws.cell(row=index + 2, column=column_mapping["CEP"]).value = info["cep"]
            ws.cell(row=index + 2, column=column_mapping["DESCRIÇÃO MATRIZ FILIAL"]).value = info["identificador_matriz_filial"]
            ws.cell(row=index + 2, column=column_mapping["TELEFONE + DDD"]).value = info["ddd_telefone_1"]
            ws.cell(row=index + 2, column=column_mapping["E-MAIL"]).value = info["email"]
            
            # Access the cell in the status column (column 17) and concatenate the API message with the existing one
            status_column = 17
            status_cell = ws.cell(row=index + 2, column=status_column)
            current_status = status_cell.value if status_cell.value else ""
            
            if status_api:
                if current_status:
                    new_status = f"{current_status}; {status_api}"
                else:
                    new_status = status_api
                status_cell.value = new_status
            msg_cnpj = f"Dados do CNPJ {cnpj} preenchidos com sucesso"
            logger_client.info(msg_cnpj)
            logger_dev.info(msg_cnpj)
        
        wb.save(output_sheet)
        msg_sheet ="Planilha preenchida com sucesso!"
        logger_client.info(msg_sheet)
        logger_dev.info(msg_sheet)
    
    except Exception as e:
        # Log the error and call the error handling function
        logger_dev.error(f"Erro ao preencher a planilha: {e}")
        handle_error("Preenchimento da planilha", "Error no preenchimento da tabela", logger_client, logger_dev)



