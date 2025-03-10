from botcity.web import WebBot, Browser, By
from botcity.maestro import *
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
import pandas as pd
import logging
import sys
import os

from utils.error_handling import handle_error

# Add the root folder to sys.path
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

from config import *
from src.utils.setup_logs import *

# Static data to execute jadlog_quote
cep_origem = 38182428
valor_coleta = '50,00'

def open_jadlog_site(bot, logger_client, logger_dev):
    # Open the jadlog quote simulation website
        bot.headless = False
        log_message = "Abre site da Jadlog no navegador"
        logger_client.info(log_message), logger_dev.info(log_message)
        bot.browse(URL_JADLOG)
        bot.maximize_window()

def jadlog_quote(output_sheet, bot, logger_client, logger_dev):
    '''
    Simulate a shipping quote on the Jadlog website using provided data.

    This function opens the Jadlog quote simulation website, fills out the required fields
    in the form with the given data, selects options from dropdowns, and clicks the 'Simular'
    button to obtain a quote. The resulting quote value is then logged.
    '''
    
    try_count = 1
    
    while try_count <=3:
        try:
            # Open output sheet
            wb = load_workbook(output_sheet)
            ws = wb.active
            
            # xlsx file reading
            df = pd.read_excel(output_sheet)
            
            # Fill the form with xlsx data
            log_message = "Iniciar cotações no side da transportadora Jadlog"
            logger_client.info(log_message), logger_dev.info(log_message)
                
            # Loop to travel dataframe cells
            for index, row in df.iterrows():
                
                # Checks empty cells and ignores quotation for the table line
                if row['CEP'] == 'Não informado':
                    log_message = f'Pulando cotação da linha {row.name + 2}. CEP do destino não declarado'
                    logger_client.warning(log_message), logger_dev.warning(log_message)
                    continue
                
                if pd.isnull(row['VALOR DO PEDIDO']):
                    log_message = f'Pulando cotação da linha {row.name + 2}. Valor do pedido não declarado'
                    logger_client.warning(log_message), logger_dev.warning(log_message)
                    continue

                if pd.isnull(row['DIMENSÕES CAIXA']):
                    log_message = f'Pulando cotação da linha {row.name + 2}. Dimensões da caixa não declaradas'
                    logger_client.warning(log_message), logger_dev.warning(log_message)
                    continue      
    
                if pd.isnull(row['PESO DO PRODUTO']):
                    log_message = f'Pulando cotação da linha {row.name + 2}. Peso não declarado'
                    logger_client.warning(log_message), logger_dev.warning(log_message)
                    continue
                
                # Separates the dimensions of the box into 3 variables
                largura, altura, comprimento = map(int, row['DIMENSÕES CAIXA'].split(' x '))
                
                # Fill in the form for quotation
                bot.find_element("//input[@id='origem']", By.XPATH).send_keys(cep_origem)
                
                bot.find_element("//input[@id='destino']", By.XPATH).clear()
                bot.find_element("//input[@id='destino']", By.XPATH).send_keys(row['CEP'])
                
                modalidade_dropdown = bot.find_element("//select[@id='modalidade']", By.XPATH)
                select = Select(modalidade_dropdown)
                if row['TIPO DE SERVIÇO JADLOG'] == 'JADLOG Expresso':
                    select.select_by_visible_text('JadLog Expresso   ')
                elif row['TIPO DE SERVIÇO JADLOG'] == 'JADLOG Econômico':
                    select.select_by_visible_text('JadLog Econômico')  
                        
                bot.find_element("//input[@id='peso']", By.XPATH).clear()
                bot.find_element("//input[@id='peso']", By.XPATH).send_keys(str(row['PESO DO PRODUTO']).replace('.',','))

                bot.find_element("//input[@id='valor_mercadoria']", By.XPATH).clear()
                bot.find_element("//input[@id='valor_mercadoria']", By.XPATH).send_keys(str(row['VALOR DO PEDIDO']).replace('.',','))
                
                bot.find_element("//input[@id='valor_coleta']", By.XPATH).clear()
                bot.find_element("//input[@id='valor_coleta']", By.XPATH).send_keys(valor_coleta)
                
                bot.find_element("//input[@id='valLargura']", By.XPATH).clear()
                bot.find_element("//input[@id='valLargura']", By.XPATH).send_keys(largura)
                
                bot.find_element("//input[@id='valAltura']", By.XPATH).clear()
                bot.find_element("//input[@id='valAltura']", By.XPATH).send_keys(altura)
                
                bot.find_element("//input[@id='valComprimento']", By.XPATH).clear()
                bot.find_element("//input[@id='valComprimento']", By.XPATH).send_keys(comprimento)
                
                # Click the 'Simular' button
                bot.wait(1000)
                bot.find_element("//input[@value='Simular']", By.XPATH).click()

                # Get the quote value
                bot.wait(3000)
                resultado = bot.find_element('//*[@id="j_idt45_content"]/span', By.XPATH).text.replace('R$ ','')
                resultado_float = float(resultado.replace(',','.'))

                # Attributes the value of the quote to the cell in the sheet
                ws[f'N{row.name + 2}'] = resultado_float
                wb.save(output_sheet)
                try_count = 5 # Break the loop if the quote succeeded

            log_message = 'Cotações realizadas com sucesso'
            logger_client.info(log_message), logger_dev.info(log_message)
            
        
        except Exception as e:
            # If an error occurs, log the error and try again
            logger_dev.error(f"{try_count}ª tentativa - Erro ao preencher formulário de cotação: {e}") 
            if try_count <3:
                logger_client.info('Ocorreu um erro na cotação. Reiniciando cotação Jadlog')
            try_count += 1
            bot.refresh()
            # Call handle_error function to handle exceptions
            handle_error("Cotação Jadlog", "Erro ao realizar cotação na Jadlog", logger_client, logger_dev)
