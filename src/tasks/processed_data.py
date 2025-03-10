import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
import os
import logging

from src.utils.error_handling import handle_error

def create_output_sheet(): 
    """
    Creates an Excel sheet with predefined headers and saves it to a specified directory.
    Applies conditional formatting to compare values in columns N and O, highlighting the cell with the smaller value.
    The file is named with the current date and time and saved as an .xlsx file.
    """
    try:
        
        # Sheet headers
        headers = [
            "CNPJ", "RAZÃO SOCIAL", "NOME FANTASIA", "ENDEREÇO", "CEP",
            "DESCRIÇÃO MATRIZ FILIAL", "TELEFONE + DDD", "E-MAIL", 
            "VALOR DO PEDIDO", "DIMENSÕES CAIXA", "PESO DO PRODUTO",
            "TIPO DE SERVIÇO JADLOG", "TIPO DE SERVIÇO CORREIOS",
            "VALOR COTAÇÃO JADLOG", "VALOR COTAÇÃO CORREIOS",
            "PRAZO DE ENTREGA CORREIOS", "Status"
        ]

        # File name with current date and time
        now = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
        file_name = f"output_sheet_{now}.xlsx"

        # Path where the file will be saved
        output_path = r"C:\customer_automation_process\src\utils\data\processed"
        
        # Make sure the path exists
        os.makedirs(output_path, exist_ok=True)

        # Full file path
        file_path = os.path.join(output_path, file_name)
        
        # Create empty dataframe with the headers
        df = pd.DataFrame(columns=headers)

        logging.info("Criando planilha de saída")

        # Save dataframe to an excel file
        df.to_excel(file_path, index=False)
        
        # Open the spreadsheet for editing
        wb = load_workbook(file_path)
        ws = wb.active

        # Define the green fill for the conditional formatting
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        # Apply conditional formatting to the range N2:N100
        ws.conditional_formatting.add('N2:N1000', CellIsRule(
            operator='lessThan', formula=['O2:O1000'], stopIfTrue=True, fill=green_fill))
        
        # Apply conditional formatting to the range O2:100
        ws.conditional_formatting.add('O2:O20', CellIsRule(
            operator='lessThan', formula=['N2:N20'], stopIfTrue=True, fill=green_fill))
            
        # Save changes in the spreadsheet
        wb.save(file_path)
        
        # Assigns the full name (with path) of the output sheet to the output_sheet variable
        # global output_sheet
        output_sheet = file_path
        logging.info(f'Planilha de saída criada com sucesso: {output_sheet}')
        
        return output_sheet
    
    except Exception as e:
        # Log the error and call the error handling function
        logging.info(f"Erro ao criar planilha de saída: {e}")
        # handle_error("Processamento de dados", "Error ao criar a tabela de saída", logger_client, logger_dev)

        
def fill_data_b4_rpachallenge(output_sheet):
    '''
    Fills empty cells from 'NOME FANTASIA' and  'TELEFONE + DDD' columns
    '''
    try:
        wb = load_workbook(output_sheet)
        ws = wb.active
        
        # xlsx file reading
        df = pd.read_excel(output_sheet)
        
        # Loop to travel dataframe cells and fill empty cells
        for index, row in df.iterrows():
            if pd.isnull(row['NOME FANTASIA']):
                ws[f'C{row.name + 2}'] = 'N/A'
            if pd.isnull(row['TELEFONE + DDD']):
                ws[f'G{row.name + 2}'] = 'Não informado'            
        
        wb.save(output_sheet)
        
    except Exception as e:
        #Log the error and call the error handling function
        logging.info(f"Erro ao preencher células em branco na planilha de saída: {e}") 
        # handle_error("Processamento de dados", "Error ao criar a tabela de saída", logger_client, logger_dev)
  
        
def fill_missing_values(output_sheet):
    '''
    Fills empty cells from the 'VALOR COTAÇÃO JADOLOG' and 'Status' columns
    '''
    try:
        wb = load_workbook(output_sheet)
        ws = wb.active
        
        # xlsx file reading
        df = pd.read_excel(output_sheet)
        
        # Loop to travel dataframe cells and fill empty cells
        for index, row in df.iterrows():
            if pd.isnull(row['VALOR COTAÇÃO JADLOG']):
                ws[f'N{row.name + 2}'] = 'N/A' 
            if pd.isnull(row['Status']):
                ws[f'Q{row.name + 2}'] = 'OK' 
            # if pd.isnull(row['VALOR COTAÇÃO CORREIOS']):
            #     ws[f'O{row.name + 2}'] = 'N/A' 
            # if pd.isnull(row['PRAZO DE ENTREGA CORREIOS']):
            #     ws[f'P{row.name + 2}'] = 'N/A'           

        wb.save(output_sheet)
        
    except Exception as e:
        # Log the error and call the error handling function
        logging.warning(f"Erro ao preencher células em branco na planilha de saída: {e}")
        # handle_error("Processamento de dados", "Preenchimento dos campos perdidos", logger_client, logger_dev)
   