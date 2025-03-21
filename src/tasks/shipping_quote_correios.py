from botcity.web import WebBot, By
import re
import logging
from config import *
from openpyxl import load_workbook
from datetime import datetime, timedelta
from src.utils.manipulate_spreadsheet import *
from src.utils.error_handling import *


def open_correios_site(bot, logger_client, logger_dev):
  """
  Open the post office website in the browser.
  """
  try:
    bot.headless = False
    bot.browse(URL_CORREIOS)
    bot.maximize_window()

    msg_open_site_correios = "Abre site dos correios no navegador."
    logger_client.info(msg_open_site_correios), logger_dev.info(msg_open_site_correios)

  except Exception as e:
     logger_dev.error(f"Erro ao abrir site no navegador. {e}")


def handle_delivery_time(delivery_text):
  """ 
  Extract business days (text) and calculate deadline
  """
  delivery_working_days = re.search(r'(\d+)\s+dias\s+úteis', delivery_text)
  if not delivery_working_days:
     return "Prazo inválido"
  
  working_days = int(delivery_working_days.group(1))
  current_date = datetime.today()

  while working_days > 0:
        current_date += timedelta(days=1)
        if current_date.weekday() < 5:  #segunda - sexta (dias uteis)
            working_days -= 1

  return current_date.strftime("%d/%m/%Y")


def fill_correios_form(bot, dest_zip, service, height, width, length, weight, row, output_sheet, logger_client, logger_dev):
  """ 
  Fill out the Correios shipping quote form and capture the quote values
  """
  try:
    #fill length cannot be less than 13
    if length < 13:
       wb = load_workbook(output_sheet)
       ws = wb.active
       ws.cell(row=row, column=17, value="Erro ao realizar a cotação Correios")
       ws.cell(row=row, column=15, value="N/A")
       ws.cell(row=row, column=16, value="N/A")
       wb.save(output_sheet)
       return 
       
    origin_zip = 38182428

    #fill in zip code fields (origin and destination)
    bot.find_element(selector="cepOrigem", by=By.NAME).send_keys(origin_zip)
    bot.find_element(selector="cepDestino", by=By.NAME).send_keys(dest_zip)

    #select service field ("PAC or "SEDEX")
    bot.find_element(selector="servico", by=By.NAME).click()
    if service == "PAC":
      bot.find_element(selector="option[value='04510']", by=By.CSS_SELECTOR).click()
    if service == "SEDEX":
      bot.find_element(selector="option[value='04014']", by=By.CSS_SELECTOR).click()
    
    #select type of packaging
    bot.find_element(selector="option[value='outraEmbalagem1']", by=By.CSS_SELECTOR).click()

    #fill in dimensions fields
    bot.find_element(selector="Altura", by=By.NAME).send_keys(height)
    bot.find_element(selector="Largura", by=By.NAME).send_keys(width)
    bot.find_element(selector="Comprimento", by=By.NAME).send_keys(length)

    #select weight
    bot.find_element(selector="peso", by=By.NAME).click()
    bot.find_element(selector=f"option[value='{weight}']:not([disabled])", by=By.CSS_SELECTOR).click()

    #Click on calculate shipping
    bot.find_element(selector="Calcular", by=By.NAME).click()

    open_tabs = bot.get_tabs()

    if len(open_tabs) > 1:
      bot.activate_tab(open_tabs[-1])

    #capture quote value
    quote_value = bot.find_element(selector="//th[contains(text(),'Valor total')]/following-sibling::td", by=By.XPATH).text
    quote_value = float(quote_value.replace('R$', '').replace(',', '.'))
    quote_value = round(quote_value, 2)

    #capture delivery deadline
    delivery = bot.find_element(selector="//th[contains(text(),'Prazo de entrega')]/following-sibling::td", by=By.XPATH).text
    delivery_date = handle_delivery_time(delivery)

    save_quote_and_delivery(output_sheet, row, quote_value, delivery_date, logger_client, logger_dev)

    close_tabs(bot)

  except Exception as e:
    logger_dev.error(f"Erro ao preencher cotação dos correios: {e}")


def save_quote_and_delivery(output_sheet, row, quote_value, delivery_date, logger_client, logger_dev):
  """
  Saves the quote value and delivery time in the output spreadsheet.
  """
  try:
    wb = load_workbook(output_sheet)
    ws = wb.active
    ws.cell(row=row, column=15, value=quote_value)
    ws.cell(row=row, column=16, value=delivery_date)
    wb.save(output_sheet)

    msg_save_quote = f"Valor da cotação e prazo de entrega salvos na linha {row}"
    logger_client.info(msg_save_quote), logger_dev.info(msg_save_quote)
   
  except Exception as e:
     # Log the error and call the error handling function
     logger_dev.error(f"Erro ao salvar cotação e prazo de entrega: {e}")
     handle_error("Cotação dos correios", "Cotação dos correios", logger_client, logger_dev)
   

def close_tabs(bot):
  """
  Close the second tab and return to the first tab.
  """
  open_tabs = bot.get_tabs()
  if len(open_tabs) > 1:
    bot.activate_tab(open_tabs[-1])
    bot.close_page()
    bot.activate_tab(open_tabs[0])
    bot.refresh()


def processed_output_sheet_quote_correios(bot, output_sheet, logger_client, logger_dev):
  """
  Processes an output spreadsheet, validates data and performs quotations.
  """
  try:
    wb = load_workbook(output_sheet)
    ws = wb.active

    msg_quote_correios = "Inicia busca de cotação dos Correios."
    logger_client.info(msg_quote_correios), logger_dev.info(msg_quote_correios)

    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        dest_zip = row[4]
        value_order = row[8]
        dimensions = row[9]
        weight = row[10]
        service = row[12]

        if not dest_zip or dest_zip == "Não informado":
          save_quote_and_delivery(output_sheet, row_index, "N/A", "N/A", logger_client, logger_dev)
          msg_empty_cep = f"CEP de destino não informado na linha {row_index}. Buscar próxima cotação."
          logger_client.warning(msg_empty_cep), logger_dev.warning(msg_empty_cep)        
          continue

        if not value_order:
          save_quote_and_delivery(output_sheet, row_index, "N/A", "N/A", logger_client, logger_dev)
          msg_empty_order = f"Valor do Pedido ausente na linha {row_index}. Buscar próxima cotação."
          logger_client.warning(msg_empty_order), logger_dev.warning(msg_empty_order)
          continue

        if not weight:
          save_quote_and_delivery(output_sheet, row_index, "N/A", "N/A", logger_client, logger_dev)
          msg_empty_quote_delivery = f"Valor de Peso ausente na linha {row_index}. Buscar próxima cotação."
          logger_client.warning(msg_empty_quote_delivery), logger_dev.warning(msg_empty_quote_delivery)
          continue

        if not dimensions or not isinstance(dimensions, str) or "x" not in dimensions:
          save_quote_and_delivery(output_sheet, row_index, "N/A", "N/A", logger_client, logger_dev)
          msg_empty_dimensions = f"Dimensões inválidas ou ausentes na linha {row_index}. Buscar próxima cotação."
          logger_client.warning(msg_empty_dimensions), logger_dev.warning(msg_empty_dimensions)
          continue

        try:
           height, width, length = map(int, dimensions.split(" x "))
           fill_correios_form(bot, dest_zip, service, height, width, length, weight, row_index, output_sheet, logger_client, logger_dev)
        except Exception as e:
          logger_dev.error(f"Erro ao converter dimensões na linha {row_index}: {e}.")
          continue

    msg_ends_quote_correios = "Finaliza busca de cotação dos Correios."
    logger_client.info(msg_ends_quote_correios), logger_dev.info(msg_ends_quote_correios)
            
  except Exception as e:
      # Log the error and call the error handling function
      logger_dev.error(f"Erro ao ler a planilha de saída: {e}")
      handle_error("Cotação dos correios", "Preenchimento da tabela com a cotação dos correiros", logger_client, logger_dev)


